import subprocess
import time
import schedule
from datetime import datetime, timedelta
import pytz

def generate_report():
    try:
        # Run the existing Python file to generate the report
        import win32com.client
        import sys
        import subprocess
        import time
        import os
        import glob
        from datetime import date, datetime, timedelta
        import pandas as pd
        from pandas import ExcelWriter
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        import numpy as np
        import xlsxwriter
        import openpyxl
        import ctypes
        from functools import reduce

        start_time = time.time()

        #Login to SAP
        subprocess.check_call([])

        path = r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"
        subprocess.Popen(path)
        time.sleep(15)

        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        connection = application.Children(application.connections.count - 1)
        session = connection.Children(0)

        #Relevant dates
        today = date.today()
        yesterday = (date.today() - timedelta(days=1)).strftime("%m/%d/%Y") #format to look dates in SAP format
        yesterday1 = (date.today() - timedelta(days=1)).strftime("%m-%d-%Y") #different date format for file saving
        half_quarter = "04/01/2022" #update
        last_quarter = "07/01/2022" #update
        end_of_fy = "09/30/2023" #update
            
        #Repository for open order report
        #folderprod = "W:\\Planning\\Production\\"
        folderdir = "C:\\Users\\"
        folderdir_soe = "C:\\Users\\\\"
        

        #File names
        test1_file = "test1.txt"
        test2_file = "test2.txt"
        test3_file = "test3.txt"
        dlv_status_file = "delivery order status.txt"
        so_status_file = "sales order status.txt"
        incompletes_file = "Incompletes.txt"
        credithold_file = "Credit hold.txt"
        shipments_file = "Shipments.txt"
        coois_file = "Coois.txt"
        saplookups = "SAP Lookups.xlsx"
        cooislookup = "COOIS.xlsx"
        inventory_file = "Inventory.txt"
        sku_by_mrpc_file = "sku by mrpc.txt"
        prodh_file = "product_hierarchy1.csv"
        customer_master_file = "customer master DF.xlsx"
        open_orders_na = "OpenOrdersNA.csv"
        order_status_history = "Order Status History.xlsx"
        grandfinale_file = "test.txt"

        print("Pulling QC Hold...")

        #Pull QC HOLD
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = ("Z_QSTOCK")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/chkPMITB").selected = True
        session.findById("wnd[0]/usr/ctxtS1_LGNUM").text = ("106")
        session.findById("wnd[0]/usr/ctxtBESTQ-LOW").text = ("")
        session.findById("wnd[0]/usr/ctxtP_VARI").text = ("/INVENTORY2")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/tbar[1]/btn[9]").press()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = 
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/usr/ctxtS1_LGNUM").text = ("105")
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/tbar[1]/btn[9]").press()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = 
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()

        print("Pulling test files...")

        #Get order list
        session.findById("wnd[0]/tbar[0]/okcd").text = ("zse16n")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/ctxtGD-TAB").text = ("vbbe")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]").sendVKey (6)
        session.findById("wnd[1]/usr/ctxtGS_SE16N_LT-NAME").text = ("order list")
        session.findById("wnd[1]/usr/txtGS_SE16N_LT-UNAME").text = ("")
        session.findById("wnd[1]").sendVKey (0)
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").Text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = test1_file
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/usr/ctxtGD-TAB").text = ("ZUPSD_DAILY_ACT")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]").sendVKey (6)
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").Text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = test2_file
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/okcd").text = ("ZUPP2P_PLANSHIP_AMT")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/ctxtSO_VSTEL-LOW").text = ("US01")
        session.findById("wnd[0]/usr/ctxtSO_VSTEL-HIGH").text = ("US02")
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").currentCellColumn = ("TOTAL")
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").selectedRows = ("0")
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").doubleClickCurrentCell()
        session.findById("wnd[0]/tbar[1]/btn[33]").press()
        session.findById("wnd[1]/tbar[0]/btn[84]").press()
        session.findById("wnd[1]/tbar[0]/btn[71]").press()
        session.findById("wnd[2]/usr/txtRSYSF-STRING").text = ("/orderstatus")
        session.findById("wnd[2]/tbar[0]/btn[0]").press()
        session.findById("wnd[3]/usr/lbl[1,2]").setFocus()
        session.findById("wnd[3]/usr/lbl[1,2]").caretPosition = 0
        session.findById("wnd[3]").sendVKey (0)
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[0]/tbar[1]/btn[45]").press()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = test3_file
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()


        print("Test files downloaded.")

        print("Waiting...")
        time.sleep(15)

        print("Getting order list...")

        #Read and clean reports
        test1clean = pd.read_csv(folderdir + test1_file, sep='\t', skiprows=3, usecols=['Sales Doc.'], dtype=str)
        test1clean.dropna()
        test2clean = pd.read_csv(folderdir + test2_file, sep='\t', skiprows=3, usecols=['Sales Doc.'], dtype=str)
        test2clean.dropna()
        test3clean = pd.read_csv(folderdir + test3_file, sep='\t', skiprows=1, usecols=['Ref.doc.'], dtype=str)
        test3clean.dropna()
        test3clean.rename(columns = {'Ref.doc.':'Sales Doc.'}, inplace=True)
        fullorderlist = pd.concat([test1clean, test2clean, test3clean], ignore_index=True) #append files
        fullorderlist.sort_values('Sales Doc.', inplace=True) 
        fullorderlist.drop_duplicates(subset='Sales Doc.', inplace=True)
        fullorderlist.to_csv(r"C:\Users\jlopez86\OneDrive - Johnson Controls\General\Order Status Report\Order Status Report (Creation Repository)\create full list.csv", index=False)

        #Execute DLVSTATUS and SOSTATUS query
        session.findById("wnd[0]/tbar[0]/okcd").text = ("sq00")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/ctxtRS38R-QNUM").text = ("dlvstatus")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/tbar[1]/btn[8]").press()

        fullorderlist.to_clipboard(excel=True, sep=None, index=False, header=None)

        session.findById("wnd[0]/usr/btn%_SP$00001_%_APP_%-VALU_PUSH").press()
        session.findById("wnd[1]/tbar[0]/btn[16]").press()
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/usr/ctxtSP$00002-LOW").text = ("sh")
        session.findById("wnd[0]/usr/ctxtSP$00003-LOW").text = ("1")
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = dlv_status_file
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/usr/ctxtRS38R-QNUM").text = ("sostatus")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/tbar[1]/btn[8]").press()

        fullorderlist.to_clipboard(excel=True, sep=None, index=False, header=None)

        session.findById("wnd[0]/usr/btn%_SP$00001_%_APP_%-VALU_PUSH").press()
        session.findById("wnd[1]/tbar[0]/btn[16]").press()
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/usr/ctxtSP$00002-LOW").text = ("sh")
        session.findById("wnd[0]/usr/ctxtSP$00003-LOW").text = ("1")
        session.findById("wnd[0]/usr/ctxtSP$00004-LOW").setFocus()
        session.findById("wnd[0]/usr/ctxtSP$00004-LOW").caretPosition = 0
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = so_status_file
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 18
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()

        print("Order and delivery report completed")
        print("Start report modeling...")

        # Create shipping conditions status definition table
        shippingconditions = pd.DataFrame({
            "SC": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", ""],
            "Order Note": ["Standard", "Pick up", "Express Delivery", "Ship Complete", "Fast Track", "Leadtime 90",
                        "Leadtime 60", "Leadtime Surcharge", "Johnstone Trade Show", "Special", "Standard"]
        })

        so_custom_column_names = ["Column1","Column2","Column3","Column4","Column5","Column6","Column7","Column8","Column9","Column10","Column11","Column12","Column13","Column14","Column15","Column16","Column17","Column18","Column19","Column20","Column21","Column22","Column23","Column24","Column25","Column26","Column27","Column28","Column29","Column30","Column31","Column32","Column33","Column34","Column35","Column36","Column37"]
        SO_Status = pd.read_csv(folderdir+so_status_file, sep="\t", skiprows=7, low_memory=False, names=so_custom_column_names, dtype=str)

        SO_Status_offset_rows1 = SO_Status.dropna(subset=['Column36']).loc[SO_Status['Column36'].str.strip() != '']
        SO_Status_offset_rows1 = SO_Status_offset_rows1[SO_Status_offset_rows1['Column37'].isnull()]
        SO_Status_offset_rows1 = SO_Status_offset_rows1.drop(columns=["Column1","Column2","Column4","Column6","Column8","Column37"])

        #SO_Status_offset_rows2 = SO_Status.dropna(subset=['Column35']).loc[SO_Status['Column35'].str.strip() != '']
        #SO_Status_offset_rows2 = SO_Status_offset_rows2.drop(columns=["Column1","Column2","Column4","Column7","Column36","Column37"])

        SO_Status_offset_rows3 = SO_Status.dropna(subset=['Column37']).loc[SO_Status['Column37'].str.strip() != '']
        SO_Status_offset_rows3 = SO_Status_offset_rows3.drop(columns=["Column1","Column2","Column4","Column7","Column20","Column21"])

        SO_Status = SO_Status.drop(columns=["Column1","Column2","Column4","Column7","Column36","Column37"])
        SO_Status.columns = ['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Rj', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'Order_Qty', 'Net_Value', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Block2', 'Group', 'SC', 'OS', 'Promise Date']

        SO_Status_offset_rows1.columns = ['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Rj', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'Order_Qty', 'Net_Value', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Block2', 'Group', 'SC', 'OS', 'Promise Date']
        #SO_Status_offset_rows2.columns = ['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Rj', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'Order_Qty', 'Net_Value', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Block2', 'Group', 'SC', 'OS', 'Promise Date']
        SO_Status_offset_rows3.columns = ['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Rj', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'Order_Qty', 'Net_Value', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Block2', 'Group', 'SC', 'OS', 'Promise Date']

        SO_Status = pd.concat([SO_Status_offset_rows1, SO_Status_offset_rows3, SO_Status], axis=0)
        SO_Status.reset_index(drop=True, inplace=True)
        SO_Status = SO_Status[(SO_Status["ABC"] != "V") | (SO_Status["OS"] != "C")]
        SO_Status = SO_Status.fillna('')
        SO_Status = SO_Status.rename(columns=lambda x: x.strip())  
        SO_Status = SO_Status.merge(shippingconditions, left_on="SC", right_on="SC", how="left")

        # Filter rows and format columns
        SO_Status = SO_Status.drop(columns=["SC", "OS"])
        SO_Status['Order Note'] = SO_Status['Order Note'].astype(str)
        SO_Status = SO_Status.drop_duplicates(subset=['Sales_Order', 'Item'], keep='first')
        SO_Status['Item'] = SO_Status['Item'].str.lstrip('0')
        SO_Status['detail'] = SO_Status['Sales_Order'] + SO_Status['Item']
        SO_Status = SO_Status[SO_Status['detail'] != ""]
        SO_Status['Promise Date'] = SO_Status.apply(lambda row: row['RDD'] if row['Promise Date'] == '00/00/0000' else row['Promise Date'], axis=1)
        SO_Status['Order_Qty'] = SO_Status['Order_Qty'].str.replace(',', '').astype(float)
        SO_Status['Net_Value'] = SO_Status['Net_Value'].str.replace(',', '').astype(float)
        SO_Status['Created_On'] = pd.to_datetime(SO_Status['Created_On'], format="%m/%d/%Y")
        SO_Status['RDD'] = pd.to_datetime(SO_Status['RDD'], format="%m/%d/%Y")
        SO_Status.loc[SO_Status['Hierarchy'] == "01111201", 'ABC'] = "C"
        SO_Status = SO_Status[SO_Status['Material'] != "INVALIDMAT"]
        SO_Status = SO_Status[SO_Status['Created_On'] > '2018-01-01']

        dlv_custom_column_names = ["Column1","Column2","Column3","Column4","Column5","Column6","Column7","Column8","Column9","Column10","Column11","Column12","Column13","Column14","Column15","Column16","Column17","Column18","Column19","Column20","Column21","Column22","Column23","Column24","Column25","Column26","Column27","Column28","Column29","Column30","Column31","Column32","Column33","Column34","Column35","Column36","Column37","Column38","Column39"]

        DLV_Status = pd.read_csv(folderdir+dlv_status_file, sep="\t", skiprows=7, low_memory=False, names=dlv_custom_column_names, dtype=str)
        DLV_Status_offset_rows1 = DLV_Status.dropna(subset=['Column38']).loc[DLV_Status['Column38'].str.strip() != '']
        DLV_Status_offset_rows1 = DLV_Status_offset_rows1[DLV_Status_offset_rows1['Column39'].isnull()]
        DLV_Status_offset_rows1 = DLV_Status_offset_rows1.drop(columns=["Column1","Column2","Column4","Column6","Column8","Column39"])

        #DLV_Status_offset_rows2 = DLV_Status.dropna(subset=['Column37']).loc[DLV_Status['Column37'].str.strip() != '']
        #DLV_Status_offset_rows2 = DLV_Status_offset_rows2.drop(columns=["Column1","Column2","Column4","Column6","Column38"])

        DLV_Status_offset_rows3 = DLV_Status.dropna(subset=['Column39']).loc[DLV_Status['Column39'].str.strip() != '']
        DLV_Status_offset_rows3 = DLV_Status_offset_rows3.drop(columns=["Column1","Column2","Column4","Column7","Column20","Column21"])

        DLV_Status = DLV_Status.drop(columns=["Column1","Column2","Column4","Column7","Column38","Column39"])
        DLV_Status.columns = ['Sales_Order','Item','PO','Hierarchy','ST','Plant','Created_On','Brand','Price_Group','Rj','Block','ABC','AV','MS','City','State','Country','Zip','Material','RDD','JS','coordinated','Sold-to','Item_Type','Delivery','DItem','DLV_Qty','Type','Ship_Date','SPI','Group','SC','Promise Date']

        DLV_Status_offset_rows1.columns = ['Sales_Order','Item','PO','Hierarchy','ST','Plant','Created_On','Brand','Price_Group','Rj','Block','ABC','AV','MS','City','State','Country','Zip','Material','RDD','JS','coordinated','Sold-to','Item_Type','Delivery','DItem','DLV_Qty','Type','Ship_Date','SPI','Group','SC','Promise Date']
        #DLV_Status_offset_rows2.columns = ['Sales_Order','Item','PO','Hierarchy','ST','Plant','Created_On','Brand','Price_Group','Rj','Block','ABC','AV','MS','City','State','Country','Zip','Material','RDD','JS','coordinated','Sold-to','Item_Type','Delivery','DItem','DLV_Qty','Type','Ship_Date','SPI','Group','SC','Promise Date']
        DLV_Status_offset_rows3.columns = ['Sales_Order','Item','PO','Hierarchy','ST','Plant','Created_On','Brand','Price_Group','Rj','Block','ABC','AV','MS','City','State','Country','Zip','Material','RDD','JS','coordinated','Sold-to','Item_Type','Delivery','DItem','DLV_Qty','Type','Ship_Date','SPI','Group','SC','Promise Date']

        DLV_Status = pd.concat([DLV_Status_offset_rows3, DLV_Status], axis=0)
        DLV_Status.reset_index(drop=True, inplace=True)

        DLV_Status = DLV_Status.fillna('')
        DLV_Status = DLV_Status.rename(columns=lambda x: x.strip())  # Remove leading/trailing whitespaces in column names.
        # Left join with shipping conditions
        DLV_Status = DLV_Status.merge(shippingconditions, left_on="SC", right_on="SC", how="left")
        DLV_Status = DLV_Status.drop(columns="SC")

        # Filter rows and format columns
        #DLV_Status.columns = ['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Rj', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Delivery', 'DItem', 'DLV_Qty', 'Type', 'Ship_Date', 'SPI', 'Group', 'Promise Date','Order Note']
        DLV_Status['Order Note'] = DLV_Status['Order Note'].astype(str)
        DLV_Status = DLV_Status.drop_duplicates(subset=['Sales_Order', 'Item', 'Delivery', 'DItem'], keep='first')
        DLV_Status['Item'] = DLV_Status['Item'].str.lstrip('0')
        DLV_Status['DItem'] = DLV_Status['DItem'].str.lstrip('0')
        DLV_Status['Promise Date'] = DLV_Status.apply(lambda row: row['RDD'] if row['Promise Date'] == '00/00/0000' else row['Promise Date'], axis=1)
        DLV_Status['detail'] = DLV_Status['Sales_Order'] + DLV_Status['Item']
        DLV_Status = DLV_Status[DLV_Status['detail'] != ""]
        DLV_Status['DLV_Qty'] = DLV_Status['DLV_Qty'].astype(str).str.replace(',', '').astype(float)
        DLV_Status['Created_On'] = pd.to_datetime(DLV_Status['Created_On'], format="%m/%d/%Y", errors='coerce')
        DLV_Status['RDD'] = pd.to_datetime(DLV_Status['RDD'], format="%m/%d/%Y", errors='coerce')
        DLV_Status.loc[DLV_Status['Hierarchy'] == "01111201", 'ABC'] = "C"
        DLV_Status = DLV_Status[DLV_Status['Created_On'] > '2018-01-01']

        # DLV Group & Greater Date
        Greater_SO = SO_Status[['Sales_Order', 'detail', 'Created_On', 'ABC', 'RDD', 'Group']]
        Greater_DLV = DLV_Status[['Sales_Order', 'detail', 'Created_On', 'ABC', 'RDD', 'Group']]
        GR8Date = pd.concat([Greater_DLV, Greater_SO])
        GR8Date = GR8Date.drop_duplicates(subset=['Sales_Order', 'detail', 'Created_On', 'RDD', 'ABC', 'Group'])

        ABC = pd.DataFrame({"ABC": ["A", "B", "C", "D", "E", "G", "P", "T", "V", ""], "Lead Time": [15, 30, 45, 30, 30, 60, 15, 45, 30, 45]})

        # Join ABC DataFrame with GR8Date
        GR8Date = GR8Date.merge(ABC, on='ABC')
        GR8Date['PLT'] = GR8Date['Created_On'] + pd.to_timedelta(GR8Date['Lead Time'], unit='D')
        GR8Date['Test'] = GR8Date[['RDD', 'PLT']].max(axis=1)
        test2 = GR8Date.groupby(['Sales_Order', 'Group'])['Test'].max().reset_index()
        test2.columns = ["Sales_Order", "Group", "Greater2"]
        GR8Date = GR8Date.merge(test2, on=['Sales_Order', 'Group'])
        GR8Date['Greater'] = GR8Date.apply(lambda row: row['Test'] if row['Group'] == "000" else row['Greater2'], axis=1)
        GR8Date['Greater'] = pd.to_datetime(GR8Date['Greater'], format='%Y-%m-%d', errors='coerce')
        GR8Date = GR8Date[['detail', 'Lead Time', 'Greater']]

        # col adjust
        SO_Status = SO_Status[['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Rj', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'Order_Qty', 'Net_Value', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Block2', 'Order Note', 'Promise Date', 'detail']]
        DLV_Status = DLV_Status[['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Rj', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Delivery', 'DItem', 'DLV_Qty', 'Type', 'Ship_Date', 'SPI', 'Order Note', 'Promise Date', 'detail']]

        # Combine qty columns by sales order/line to get open qty
        sodetail = SO_Status[['detail', 'Order_Qty']]
        dlvdetail = DLV_Status[['detail', 'DLV_Qty']]
        dlvdetail = dlvdetail.groupby('detail')['DLV_Qty'].sum().reset_index()
        id = pd.merge(sodetail, dlvdetail, on='detail', how='outer')
        id = id.fillna(0)
        id['Open_Qty'] = id['Order_Qty'] - id['DLV_Qty']

        # Get order value to calculate $'s in DLV + % of order in DLV
        sosum = SO_Status[['detail', 'Sales_Order', 'Net_Value', 'Order_Qty']]
        dlvmerge = DLV_Status[['detail', 'Delivery', 'DItem', 'DLV_Qty']]
        dlvmerge = pd.merge(dlvmerge, sosum, on='detail', how='left')
        dlvmerge['DLV_Value'] = (dlvmerge['DLV_Qty'] / dlvmerge['Order_Qty']) * dlvmerge['Net_Value']
        sosum = pd.merge(sosum, dlvdetail, on='detail', how='left')
        sosum = sosum.fillna(0)
        sosum['DLV_Value'] = (sosum['DLV_Qty'] / sosum['Order_Qty']) * sosum['Net_Value']
        valuemerge = sosum
        sosum = sosum[['Sales_Order', 'Net_Value', 'Order_Qty', 'DLV_Qty', 'DLV_Value']]
        sosum = sosum.groupby('Sales_Order').agg({'Net_Value': 'sum', 'DLV_Value': 'sum'}).reset_index()
        sosum['PercentDLV'] = sosum['DLV_Value'] / sosum['Net_Value']
        sosum = sosum[['Sales_Order', 'Net_Value', 'PercentDLV']]
        sosum.columns = ["Sales_Order", "Order_Value", "%_DLV"]
        valuemerge['Open_Value'] = valuemerge['Net_Value'] - valuemerge['DLV_Value']
        valuedlv = dlvmerge[['Delivery', 'DItem', 'DLV_Value']]
        valueopen = valuemerge[['detail', 'Open_Value']]

        # Remove extra columns, rejected lines, empty lines, source 1, warranty, shipped(invoiced) lines
        id = id[['detail', 'Open_Qty']]
        statusreport = pd.merge(SO_Status, id, on='detail', how='left')
        statusreport = pd.merge(statusreport, sosum, on='Sales_Order', how='left')
        statusreport = statusreport[['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Rj', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'Net_Value', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Block2', 'Order Note', 'Promise Date', 'detail', 'Open_Qty', 'Order_Value', '%_DLV']]
        statusreport = pd.merge(statusreport, valueopen, on='detail', how='left')
        statusreport = statusreport[statusreport['Open_Qty'] != 0]
        statusreport = statusreport[statusreport['Rj'] == ""]
        statusreport = statusreport[statusreport['Item_Type'] != "ZTAX"]
        statusreport = statusreport[statusreport['Item_Type'] != "ZXWM"]
        statusreport = statusreport[statusreport['Item_Type'] != "ZXW2"]
        statusreport = statusreport[statusreport['Item_Type'] != 'ZCF1']
        statusreport = statusreport[(statusreport['Plant'] == "US01") | (statusreport['Plant'] == "US02") | (statusreport['Plant'] == "US03") | (statusreport['Plant'] == "MX25") | (statusreport['Plant'] == "")]
        DLV_Status = DLV_Status[DLV_Status['Item_Type'] != "ZXWM"]
        DLV_Status = DLV_Status[DLV_Status['Item_Type'] != "ZXW2"]
        DLV_Status = DLV_Status[DLV_Status['Item_Type'] != 'ZCF1']
        DLV_Status = DLV_Status[(DLV_Status['Plant'] == "US01") | (DLV_Status['Plant'] == "US02") | (DLV_Status['Plant'] == "US03") | (DLV_Status['Plant'] == "MX25") | (DLV_Status['Plant'] == "")]
        DLV_Status = pd.merge(DLV_Status, sosum, on='Sales_Order', how='left')
        DLV_Status = pd.merge(DLV_Status, valuedlv, left_on=['Delivery', 'DItem'], right_on=['Delivery', 'DItem'], how='left')
        DLV_Status = DLV_Status[DLV_Status['Ship_Date'] == "00/00/0000"]

        sales = statusreport[['Sales_Order']].drop_duplicates()
        delivery = DLV_Status[['Delivery']].drop_duplicates()

        with pd.ExcelWriter(
            r"",
            engine='xlsxwriter'
        ) as writer:
            sales.to_excel(writer, sheet_name='Orders', index=False)
            delivery.to_excel(writer, sheet_name='Deliveries', index=False)

        print("Pulling Order look up list...")

        #Read SAP Lookups excel file
        orderlookuplist = pd.read_excel(folderdir + saplookups, sheet_name=0)
        deliverieslookuplist = pd.read_excel(folderdir + saplookups, sheet_name=1)

        #Copy orders on clipboard
        orderlookuplist.to_clipboard(excel=True, sep=None, index=False, header=None)

        #Get material list for COOIS
        session.findById("wnd[0]/tbar[0]/okcd").text = ("zse16n")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/ctxtGD-TAB").text = ("vbuv")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]").sendVKey (6)
        session.findById("wnd[1]/usr/ctxtGS_SE16N_LT-NAME").text = ("orderstatus")
        session.findById("wnd[1]/usr/txtGS_SE16N_LT-UNAME").text = ("")
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/btnPUSH[4,1]").setFocus()
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/btnPUSH[4,1]").press()
        session.findById("wnd[1]/tbar[0]/btn[34]").press()
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = incompletes_file
        session.findById("wnd[1]/usr/ctxtDY_FILE_ENCODING").setFocus()
        session.findById("wnd[1]/usr/ctxtDY_FILE_ENCODING").caretPosition = 0
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()

        #Copy deliveries on clipboard
        deliverieslookuplist.to_clipboard(excel=True, sep=None, index=False, header=None)

        session.findById("wnd[0]/usr/ctxtGD-TAB").text = ("vbuk")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]").sendVKey (6)
        session.findById("wnd[1]/usr/ctxtGS_SE16N_LT-NAME").text = ("orderstatus")
        session.findById("wnd[1]/usr/txtGS_SE16N_LT-UNAME").text = ("")
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/btnPUSH[4,1]").setFocus()
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/btnPUSH[4,1]").press()
        session.findById("wnd[1]/tbar[0]/btn[34]").press()
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = credithold_file
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 10
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/okcd").text = ("sq00")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/ctxtRS38R-QNUM").text = ("shipmentstatus")
        session.findById("wnd[0]/usr/ctxtRS38R-QNUM").setFocus()
        session.findById("wnd[0]/usr/ctxtRS38R-QNUM").caretPosition = 14
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/btn%_SP$00001_%_APP_%-VALU_PUSH").press()
        session.findById("wnd[1]/tbar[0]/btn[16]").press()
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = shipments_file
        session.findById("wnd[1]/usr/ctxtDY_FILE_ENCODING").setFocus()
        session.findById("wnd[1]/usr/ctxtDY_FILE_ENCODING").caretPosition = 0
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()

        incompletes = pd.read_csv(folderdir + incompletes_file, sep='\t', skiprows=1, dtype=str)
        incompletes = incompletes.drop(columns=["Unnamed: 0"])
        incompletes['Item'] = incompletes['Item'].str.lstrip(' ')
        incompletes.fillna(value='', inplace=True) 
        incompletes.columns = ["Sales_Order", "Item", "Field", "Incomplete"]
        incompletes['detail'] = incompletes["Sales_Order"] + incompletes["Item"]

        Credit = pd.read_csv(folderdir+credithold_file, sep='\t', skiprows=1, dtype=str)
        Credit = Credit[['Document', 'OvCS']]
        Credit.fillna(value='', inplace=True)
        Credit.columns = ["Delivery", "OVCS"]

        Shipments = pd.read_csv(folderdir+shipments_file, sep='\t', skiprows=6, dtype=str)
        Shipments = Shipments.drop(columns=["Unnamed: 0", "Unnamed: 2"])
        Shipments.fillna(value='', inplace=True)
        Shipments.columns = ["Delivery", "Shipment", "SHP_Plant", "Created", "Ship_Type", "Carrier", "SHP_Status"]

        DLV_Status = pd.merge(DLV_Status, Shipments, on='Delivery', how='left')
        DLV_Status = pd.merge(DLV_Status, Credit, on='Delivery', how='left')
        DLV_Status['Created_On'] = pd.to_datetime(DLV_Status['Created_On'], format="%Y-%m-%d", errors='coerce')
        DLV_Status['RDD'] = pd.to_datetime(DLV_Status['RDD'], format="%Y-%m-%d", errors='coerce')

        incomergre = incompletes[['detail', 'Incomplete']].drop_duplicates()

        statusreport = pd.merge(statusreport, incomergre, on='detail', how='left')
        statusreport = statusreport[['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'Open_Qty', 'Open_Value', 'Order_Value', '%_DLV', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Block2', 'Order Note', 'Promise Date', 'Incomplete', 'detail']]
        statusreport = statusreport[statusreport['Open_Qty'] > 0]

        DLV_Status = DLV_Status[['Sales_Order', 'Item', 'PO', 'Hierarchy', 'ST', 'Plant', 'Created_On', 'Brand', 'Price_Group', 'Block', 'ABC', 'AV', 'MS', 'City', 'State', 'Country', 'Zip', 'Material', 'DLV_Qty', 'DLV_Value', 'Order_Value', '%_DLV', 'RDD', 'JS', 'coordinated', 'Sold-to', 'Item_Type', 'Delivery', 'DItem', 'Order Note', 'Promise Date', 'SPI', 'OVCS', 'Shipment', 'SHP_Plant', 'Created', 'Ship_Type', 'Carrier', 'SHP_Status', 'detail']]

        finale = pd.concat([statusreport, DLV_Status], axis=0, ignore_index=True)
        finale = finale.applymap(lambda x: '' if pd.isna(x) else x)
        finale = finale[finale['Material'] != 'S1-OBSOLETE']
        finale = pd.merge(finale, GR8Date, on='detail', how='left')
        finale = finale.drop(columns=["detail", "Lead Time"])

        model = finale[['Plant', 'Material', 'Open_Qty', 'DLV_Qty']]
        model = model.copy()
        model['Open_Qty'] = pd.to_numeric(model['Open_Qty'])
        model['DLV_Qty'] = pd.to_numeric(model['DLV_Qty'])
        model = model.fillna(0)
        model = model.groupby(['Plant', 'Material']).agg('sum').reset_index()

        normanmodel = model[model['Plant'] == 'US01']
        wichitamodel = model[model['Plant'] == 'US02']
        S1model = model[model['Plant'] == 'US03']
        C2Model = model[model['Plant'] == 'MX25']

        #Set S1 and C2 inventory in 0
        S12 = S1model.copy()
        S12['Unrestricted'] = 0
        S12['QC_HOLD'] = 0
        S12['MS'] = 'P2'
        S12 = S12[['Material', 'Unrestricted', 'QC_HOLD', 'Open_Qty', 'DLV_Qty', 'MS']]
        S12['Plant'] = 'PARTS'

        C22 = C2Model.copy()
        C22['Unrestricted'] = 0
        C22['QC_HOLD'] = 0
        C22['MS'] = 'P2'
        C22 = C22[['Material', 'Unrestricted', 'QC_HOLD', 'Open_Qty', 'DLV_Qty', 'MS']]
        C22['Plant'] = 'MX25'

        wichitainv = pd.read_csv(folderdir+wichitaqc_file, sep='\t', header=0, skiprows=3, dtype=str)
        wichitainv = wichitainv.rename(columns=lambda x: x.strip())  # Remove leading/trailing whitespaces in column names
        wichitainv = wichitainv[['Material', 'S', 'Total Stock', 'MS']]
        wichitainv = wichitainv.fillna('Unrestricted')
        wichitainv['S'] = wichitainv['S'].replace('Q', 'QC_HOLD')
        wichitainv = wichitainv[(wichitainv['S'] != 'S') & (wichitainv['S'] != 'R')]
        wichitainv['Total Stock'] = pd.to_numeric(wichitainv['Total Stock'])
        wichitainv = wichitainv[wichitainv['Total Stock'] > 0]
        wichitainv = pd.pivot_table(wichitainv, values='Total Stock', index=['Material','MS'], columns=['S'], aggfunc='sum', fill_value=0).reset_index()
        wichitainv = wichitainv[['Material', 'MS', 'Unrestricted', 'QC_HOLD']]

        normaninv = pd.read_csv(folderdir+normanqc_file, sep='\t', header=0, skiprows=3, dtype=str)
        normaninv = normaninv.rename(columns=lambda x: x.strip())  # Remove leading/trailing whitespaces in column names
        normaninv = normaninv[['Material', 'S', 'Total Stock', 'MS']]
        normaninv = normaninv.fillna('Unrestricted')
        normaninv['S'] = normaninv['S'].replace('Q', 'QC_HOLD')
        normaninv = normaninv[(normaninv['S'] != 'S') & (normaninv['S'] != 'R')]
        normaninv['Total Stock'] = pd.to_numeric(normaninv['Total Stock'])
        normaninv = normaninv[normaninv['Total Stock'] > 0]
        normaninv = pd.pivot_table(normaninv, values='Total Stock', index=['Material','MS'], columns=['S'], aggfunc='sum', fill_value=0).reset_index()
        normaninv = normaninv[['Material', 'MS', 'Unrestricted', 'QC_HOLD']]

        norman = normanmodel.merge(normaninv, on='Material', how='left')
        norman2 = normaninv.merge(normanmodel, on='Material', how='left')
        norman2 = norman2[['Material', 'Unrestricted', 'QC_HOLD', 'Open_Qty', 'DLV_Qty', 'MS']]
        norman2['Plant'] = 'US01'
        wichita = wichitamodel.merge(wichitainv, on='Material', how='left')
        wichita2 = wichitainv.merge(wichitamodel, on='Material', how='left')
        wichita2 = wichita2[['Material', 'Unrestricted', 'QC_HOLD', 'Open_Qty', 'DLV_Qty', 'MS']]
        wichita2['Plant'] = 'US02'

        inventory = pd.concat([norman, wichita, S12, C22])
        inventory = inventory.fillna(0)
        inventory = inventory[['Plant', 'Material', 'Open_Qty', 'DLV_Qty', 'Unrestricted','QC_HOLD']]
        inventory.columns = ['Plant', 'Material', 'Total_Open', 'Total_DLV', 'Unrestricted','QC_Hold']

        finale = finale.merge(inventory, left_on=['Material', 'Plant'], right_on=['Material', 'Plant'], how='left')

        prodh = pd.read_csv(folderdir_soe+inventory_file, sep='\t', skiprows=5, dtype=str)
        prodh = prodh[['Material', 'Product hierarchy']]
        prodh = prodh.drop_duplicates(subset=['Material'])

        inventory = inventory.merge(prodh, on='Material', how='left')
        inventory.to_csv(r"C:\Users\jlopez86\OneDrive - Johnson Controls\General\Order Status Report\Order Status Report (Creation Repository)\Availablility.txt", index=False)

        # Rest of the code (Setting Status values)
        today = datetime.today().date()
        finale['Status2'] = ""
        finale['Finish_Date'] = "2000-01-01"
        finale['Finish_Date'] = pd.to_datetime(finale['Finish_Date'])
        finale['Greater'] = pd.to_datetime(finale['Greater']).dt.date
        shipment = finale[finale['Shipment'] != ""].copy()
        shipment['Status'] = "Shipment Processing"
        finale = finale[finale['Shipment'] == ""]
        incomp = finale[finale['Incomplete'] != ""].copy()
        incomp['Status'] = "Incomplete"
        finale = finale[finale['Incomplete'] == ""]
        dhold = finale[finale['Block'] == "Z4"].copy()
        dhold['Status'] = "DLV Hold"
        mod1 = dhold[['Sales_Order', 'Item_Type']]
        mod1 = mod1[mod1['Item_Type'] == "ZTAC"].copy()
        mod1.rename(columns={'Item_Type': 'IT2'}, inplace=True)
        mod1 = mod1.drop_duplicates()
        dhold = dhold.merge(mod1, on='Sales_Order', how='left')
        dhold['Status'] = np.where(dhold['IT2'] == "ZTAC", "MOD Hold", "DLV Hold")
        dhold.loc[pd.isna(dhold['Status']), 'Status'] = "DLV Hold"
        dhold = dhold.drop(columns=['IT2'])
        finale = finale[finale['Block'] != "Z4"]
        dblock = finale[(finale['Block'] != "") | (finale['Block2'] != "")].copy()
        dblock['Status'] = "DLV Block"
        finale = finale[(finale['Block'] == "") & (finale['Block2'] == "")]
        credit = finale[finale['OVCS'] == "B"].copy()
        credit['Status'] = "Credit Hold"
        finale = finale[finale['OVCS'] != "B"]
        logh = finale[finale['SPI'] == "LOGH"].copy()
        logh['Status'] = "Logistics Hold"
        finale = finale[finale['SPI'] != "LOGH"]
        shpc = finale[finale['SPI'] == "SHPC"].copy()
        shpc['Status'] = "Ship Complete"
        finale = finale[finale['SPI'] != "SHPC"]
        pool = finale[(finale['ST'] == "   17") | (finale['ST'] == "   18")].copy()
        pool = pool[pool['SPI'] == "POOL"]
        pool['Status'] = "POOL"
        finale = finale[(finale['ST'] != "   17") & (finale['ST'] != "   18") | (finale['SPI'] != "POOL")]
        dlv = finale[finale['Delivery'] != ""].copy()
        dlv['Status'] = "Logistics Planning"
        finale = finale[finale['Delivery'] == ""]
        vds = finale[finale['ABC'] == "V"].copy()
        vds['Status'] = "VDS Item"
        finale = finale[finale['ABC'] != "V"]
        obsolete = finale[finale['MS'] == "S4"].copy()
        obsolete['Status'] = "Obsolete Model"
        finale = finale[finale['MS'] != "S4"]
        mod = finale[finale['Item_Type'] == "ZTAC"].copy()
        mod['Status'] = "MOD Item"
        finale = finale[finale['Item_Type'] != "ZTAC"]
        finale['phaseout'] = np.where((finale['Unrestricted'] + finale['QC_Hold']) < (finale['Total_DLV'] + finale['Total_Open']), "uhoh", "")
        phaseout = finale[(finale['MS'] == "S5") & (finale['phaseout'] == "uhoh")].copy()
        phaseout = phaseout.drop(columns=['phaseout'])
        phaseout['Status'] = "Phase Out"
        finale = finale[(finale['MS'] != "S5") | (finale['phaseout'] != "uhoh")]
        finale = finale.drop(columns=['phaseout'])
        finale['Status'] = np.where(finale['Greater'] > (today + timedelta(days=7)), "Future Order", "Awaiting Allocation")
        lateprod = finale[finale['Status'] == "Awaiting Allocation"].copy()
        finale = finale[finale['Status'] != "Awaiting Allocation"]

        coois_list = lateprod[['Material']].drop_duplicates()
        coois_list.to_excel(r"", index=False) # Change to list only for SAP pull ?

        print("Pulling COOIS list...")

        cooislookuplist = pd.read_excel(folderdir + cooislookup, sheet_name=0)
        cooislookuplist.to_clipboard(excel=True, sep=None, index=False, header=None)

        #Get COOIS file
        session.findById("wnd[0]/tbar[0]/okcd").text = ("coois")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/tbar[1]/btn[17]").press()
        session.findById("wnd[1]/usr/txtV-LOW").text = ("orderstatus")
        session.findById("wnd[1]").sendVKey (0)
        session.findById("wnd[1]/usr/txtENAME-LOW").text = ("")
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/btn%_S_MATNR_%_APP_%-VALU_PUSH").press()
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").pressToolbarButton ("&NAVIGATION_PROFILE_TOOLBAR_EXPAND")
        session.findById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = coois_file
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()


        coois = pd.read_csv(folderdir+coois_file, sep='\t', skiprows=1, dtype=str, encoding='latin1')
        coois = coois.rename(columns=lambda x: x.strip())
        coois = coois[['Order', 'Material', 'Target qty', 'Del. qty', 'Basic fin.']]
        coois[['Target qty', 'Del. qty']] = coois[['Target qty', 'Del. qty']].apply(lambda col: col.str.replace(',', '').astype(float))
        coois['Open'] = coois['Target qty'] - coois['Del. qty']
        coois = coois.drop(columns=['Target qty', 'Del. qty'])
        coois.columns = ["Prod", "Material", "Finish_Date", "Open_Qty"]
        coois['Finish_Date'] = pd.to_datetime(coois['Finish_Date'], format="%m/%d/%Y")
        coois = coois.sort_values(by=['Material', 'Finish_Date'])

        lateprod = lateprod.sort_values(by=['Material', 'Greater']).reset_index(drop=True)
        lateprod['Open_Qty'] = pd.to_numeric(lateprod['Open_Qty'])
        lateprod['Status2'] = 'test'
        lateprod['Finish_Date'] = pd.to_datetime(lateprod['Finish_Date'])
        lateprod[['Unrestricted', 'QC_Hold']] = lateprod[['Unrestricted', 'QC_Hold']].fillna(0)
        lateprod['Available'] = lateprod['Unrestricted'] - lateprod['Total_DLV']

        lateprod['Balance'] = lateprod.groupby('Material').apply(
                lambda x: x['Available'].iat[0] - (x['Open_Qty'].cumsum() - x['Open_Qty'])
                if x.name[0] == x['Material'].iloc[0]
                else x['Available'] - x['Open_Qty'].cumsum()
            ).reset_index(drop=True)
        lateprod['Status2'] = lateprod['Balance'].apply(lambda x: 'In Stock' if x >= 0 else 'test')

        lateprod['Available_QC'] = lateprod['Available'] + lateprod['QC_Hold']
        lateprod['Balance_QC'] = lateprod.groupby('Material').apply(
                lambda x: x['Available_QC'].iat[0] - (x['Open_Qty'].cumsum() - x['Open_Qty'])
                if x.name[0] == x['Material'].iloc[0]
                else x['Available_QC'] - x['Open_Qty'].cumsum()
            ).reset_index(drop=True)

        lateprod['Status2'] = lateprod.apply(lambda row: 'QC HOLD' if row['Balance_QC'] >= 0 and row['Status2'] == 'test' else row['Status2'], axis=1)
        lateprod = lateprod.drop(columns=['Available', 'Balance', 'Available_QC', 'Balance_QC'])
        complete = lateprod[(lateprod['Order Note'] == 'Ship Complete') & (lateprod['Status2'] == 'In Stock')].copy()
        complete['Status2'] = 'In Stock - Ship Complete'
        lateprod = lateprod[(lateprod['Order Note'] != 'Ship Complete') | (lateprod['Status2'] != 'In Stock')]
        lateprod = pd.concat([lateprod, complete])
        noresult = lateprod[lateprod['Status2'] == 'test']
        lateprod = lateprod[lateprod['Status2'] != 'test']

        phg_master = pd.read_csv(folderdir_soe+prodh_file)
        phg_master = phg_master[['series_id', 'line']]
        phg_master.columns = ['Hierarchy', 'PHG']
        phg_master['PHG'] = phg_master['PHG'].replace({'Predator': 'Pro', 'CDR - Pkg': 'Core'}, regex=True)
        noresult = pd.merge(noresult, phg_master, on='Hierarchy', how='left')
        conditions = [
            (noresult['PHG'] == "Coils - MH") | (noresult['PHG'] == "COM ACCY") | (noresult['PHG'] == "MH ACCY"),
            True
        ]
        choices = ['Purchased Part', 'Multiple Production Orders']
        noresult['Status2'] = np.select(conditions, choices, default='')
        noresult = noresult.drop(columns=['PHG'])
        lateprod = pd.concat([lateprod, noresult])

        grandfinale = pd.concat([incomp, dhold, dblock, credit, logh, obsolete, phaseout, shpc, pool, dlv, shipment, lateprod, finale, mod, vds])

        sku_list = grandfinale[['Material']]
        sku_list.drop_duplicates(subset='Material', inplace=True)
        sku_list.to_clipboard(excel=True, sep=None, index=False, header=None)

        session.findById("wnd[0]/tbar[0]/okcd").text = ("ZSE16N")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/ctxtGD-TAB").text = ("MARC")
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/usr/txtGD-MAX_LINES").text = ("")
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/btn").setFocus()
        session.findById("wnd[0]/tbar[1]/btn[18]").press()
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/chkGS_SELFIELDS-MARK[5,1]").selected = True
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/chkGS_SELFIELDS-MARK[5,2]").selected = True
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/chkGS_SELFIELDS-MARK[5,15]").selected = True
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/btnPUSH[4,1]").setFocus()
        session.findById("wnd[0]/usr/tblSAPLSE16NSELFIELDS_TC/btnPUSH[4,1]").press()
        session.findById("wnd[1]/tbar[0]/btn[34]").press()
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").pressToolbarContextButton ("&MB_EXPORT")
        session.findById("wnd[0]/usr/cntlRESULT_LIST/shellcont/shell").selectContextMenuItem ("&PC")
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = folderdir
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = sku_by_mrpc_file
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 11
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        session.findById("wnd[0]/tbar[0]/btn[3]").press()

        sku_by_mrpc = pd.read_csv(folderdir+sku_by_mrpc_file, sep='\t', skiprows=1, dtype=str, encoding='latin1')
        sku_by_mrpc = sku_by_mrpc.rename(columns=lambda x: x.strip())
        sku_by_mrpc = sku_by_mrpc[['Material', 'Plnt', 'MRPC']]
        sku_by_mrpc = sku_by_mrpc[sku_by_mrpc['Plnt'].isin(['1001', '1002', '1025', '1003'])]
        sku_by_mrpc['Plnt'] = sku_by_mrpc['Plnt'].replace(['1001', '1002', '1025', '1003'], ['US01', 'US02', 'MX25', 'US03'])

        customer_master = pd.read_excel(folderdir+customer_master_file, sheet_name='Customer_Master')
        customer_master = customer_master[['SAPLookup', 'cm.vlabel.2', 'Channel Final', 'RAM Final']]
        customer_master.columns = ['Sold-to', 'Name', 'Channel', 'CSR']

        grandfinale = pd.merge(grandfinale, customer_master, on='Sold-to', how='left')
        grandfinale = pd.merge(grandfinale, phg_master, on='Hierarchy', how='left')
        grandfinale = pd.merge(grandfinale, sku_by_mrpc,  how='left', left_on=['Material','Plant'], right_on = ['Material','Plnt'])

        natag = pd.read_csv(folderdir+open_orders_na, usecols=['Reference', 'SAPOrd1', 'SAPOrd2', 'SAPOrd3'], dtype='str', encoding='latin1')
        natag = pd.concat([
            natag[['Reference', 'SAPOrd1']].dropna().rename(columns={'SAPOrd1': 'Sales_Order'}),
            natag[['Reference', 'SAPOrd2']].dropna().rename(columns={'SAPOrd2': 'Sales_Order'}),
            natag[['Reference', 'SAPOrd3']].dropna().rename(columns={'SAPOrd3': 'Sales_Order'})
        ])
        natag['Sales_Order'] = natag['Sales_Order'].astype(str).str.lstrip('0')
        natag.drop_duplicates(inplace=True)
        natag.rename(columns={'Reference': 'NA TAG'}, inplace=True)

        grandfinale = pd.merge(grandfinale, natag[['NA TAG']], left_on='Sales_Order', right_on='NA TAG', how='left')
        grandfinale['Created_On'] = pd.to_datetime(grandfinale['Created_On'])
        grandfinale['RDD'] = pd.to_datetime(grandfinale['RDD'])
        grandfinale['Created'] = pd.to_datetime(grandfinale['Created'])
        grandfinale['Promise Date'] = pd.to_datetime(grandfinale['Promise Date'])
        grandfinale['Greater'] = pd.to_datetime(grandfinale['Greater'])
        grandfinale['Finish_Date'] = pd.to_datetime(grandfinale['Finish_Date'])

        condition_next15 = grandfinale['Greater'] < (pd.Timestamp('today') + pd.DateOffset(days=15))
        grandfinale['Next_15'] = np.where(condition_next15, 'Yes', 'No')
        condition_pastdues = grandfinale['RDD'] >= pd.Timestamp('today')
        grandfinale['Past Dues'] = np.where(condition_pastdues, 'N', 'Y')

        #formatting final output
        column_layout = ['Sales_Order','Item','PO','Hierarchy','PHG','ST','Plant','Created_On','Brand','Price_Group','Block','ABC','MRPC','AV','MS','City','State','Country','Zip','Material','Open_Qty','Open_Value','DLV_Qty','DLV_Value','%_DLV','Order_Value','RDD','JS','coordinated','Sold-to','Name','Channel','CSR','NA TAG','Item_Type','Block2','Order Note','Incomplete','Delivery','DItem','SPI','OVCS','Shipment','SHP_Plant','Created','Ship_Type','Carrier','SHP_Status','Greater','Promise Date','Total_Open','Total_DLV','Unrestricted','QC_Hold','Status','Status2','Finish_Date','Next_15','Past Dues']
        grandfinale = grandfinale.reindex(columns=column_layout)
        grandfinale[['Open_Qty','DLV_Qty','Total_Open','Total_DLV']] = grandfinale[['Open_Qty','DLV_Qty','Total_Open','Total_DLV']].apply(lambda col: pd.to_numeric(col, errors='coerce'))
        grandfinale[['Open_Value','DLV_Value','Order_Value']] = grandfinale[['Open_Value','DLV_Value','Order_Value']].apply(lambda col: pd.to_numeric(col, errors='coerce'))
        columns_to_fill = ['Open_Qty','Open_Value','DLV_Qty','DLV_Value','Order_Value','Total_Open','Total_DLV','Unrestricted','QC_Hold']
        grandfinale[columns_to_fill] = grandfinale[columns_to_fill].fillna(0)

        date_columns = ['Created_On', 'RDD', 'Created', 'Greater', 'Promise Date', 'Finish_Date']
        grandfinale[date_columns] = grandfinale[date_columns].apply(lambda x: pd.to_datetime(x, format='%Y-%m-%d', errors='coerce'))
        grandfinale[date_columns] = grandfinale[date_columns].apply(lambda x: x.dt.date)

        print("Model done, exporting to excel...")

        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        def save_workbook(workbook, file_path):
            workbook.save(file_path)

        timestamp = datetime.now().strftime('%Y-%m-%d')
        grandfinale_xlsname = f'Order Status Report_{timestamp}.xlsx'
        grandfinale_xlsname_current = f'Order Status Report_current.xlsx'

        # Write DataFrame to Excel file using pandas
        with pd.ExcelWriter(folderdir_desktop + grandfinale_xlsname, engine='openpyxl', date_format='m/d/yyyy') as excel_writer:
            grandfinale.to_excel(excel_writer, sheet_name='Report', index=False)

        book = load_workbook(folderdir_desktop + grandfinale_xlsname)

        worksheet = book['Report']

        header_row = worksheet[1]

        # formatting style
        header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        font_style = Font(color='000000', bold=True)
        alignment_style = Alignment(horizontal='left')

        # Apply formatting to headers
        for cell in header_row:
            cell.fill = header_fill
            cell.font = font_style
            cell.alignment = alignment_style
            cell.border = None  # Remove default border for header cells

        save_workbook(book, folderdir_outputs + grandfinale_xlsname)
        save_workbook(book, folderdir_upg + grandfinale_xlsname_current)



        # Saving past due burndown data
        import datetime

        if datetime.datetime.today().weekday() == 0:
            print("Exporting past due history file...")

            # Run the code for Fridays
            past_due_summary = grandfinale[grandfinale['Past Dues'] == 'Y'].copy()
            past_due_summary['Date'] = pd.to_datetime('today')
            past_due_summary = past_due_summary[past_due_summary['Plant'].isin(['US01', 'US02', 'MX25'])]
            plant_mapping = {'US01': 'COM', 'US02': 'RESI', 'MX25': 'RESI'}
            past_due_summary['Plant'] = past_due_summary['Plant'].replace(plant_mapping)

            past_due_summary = past_due_summary.groupby(['Date', 'Plant']).agg({
                'Open_Qty': 'sum',
                'Open_Value': 'sum',
                'DLV_Qty': 'sum',
                'DLV_Value': 'sum'
            }).reset_index(drop=False)

            past_due_summary['Total_Qty'] = past_due_summary['Open_Qty'] + past_due_summary['DLV_Qty']
            past_due_summary['Total_Value'] = past_due_summary['Open_Value'] + past_due_summary['DLV_Value']
            past_due_summary['Date'] = pd.to_datetime(past_due_summary['Date']).apply(lambda x: x.date())


            past_due_path = folderdir_historical + 'Past Due Burndown.xlsx'

            with pd.ExcelFile(past_due_path) as xls:
                past_dues_sheet = pd.read_excel(xls, 'Past Dues')

            past_dues_sheet['Date'] = pd.to_datetime(past_dues_sheet['Date']).apply(lambda x: x.date())

            combined_df_past_dues = pd.concat([past_due_summary, past_dues_sheet], ignore_index=True)
            combined_df_past_dues.to_excel(past_due_path, sheet_name='Past Dues', index=False)
        else:
            print("This code only runs on Mondays.")

        pd.options.display.float_format = '{:,.0f}'.format
        
        print("-----------------------------")
        print("Number of lines: ")
        print(grandfinale['Sales_Order'].count())

        print("-----------------------------")
        print("Unique orders: ")
        print(grandfinale['Sales_Order'].nunique())

        print("-----------------------------")
        print("Sum of open qty: ")
        print(grandfinale['Open_Qty'].sum())

        print("-----------------------------")
        print("Sum of dlv qty: ")
        print(grandfinale['DLV_Qty'].sum())


        print("-----------------------------")
        print(grandfinale.groupby('Plant')[['Open_Qty', 'DLV_Qty', 'Open_Value', 'DLV_Value']].sum())

        print("-----------------------------")
        print(grandfinale.groupby('Past Dues')[['Open_Qty', 'DLV_Qty', 'Open_Value', 'DLV_Value']].sum())
            

        print('Report generated successfully.')


        # Calculate the elapsed time
        elapsed_time = (time.time() - start_time) / 60

        # Print the elapsed time in seconds
        print(f"Elapsed time: {elapsed_time} minutes")

    except Exception as e:
        print(f'Failed to generate report: {str(e)}')

# Set the US Central Time (CT) timezone
ct_timezone = pytz.timezone('America/Monterrey')

# Schedule the report to run at 6:00 AM CT from Monday to Friday
schedule.every().monday.at('08:04').do(generate_report)
schedule.every().tuesday.at('06:00').do(generate_report)
schedule.every().wednesday.at('06:00').do(generate_report)
schedule.every().thursday.at('06:00').do(generate_report)
schedule.every().friday.at('06:00').do(generate_report)
#schedule.every().saturday.at('06:00').do(generate_report)
#schedule.every().sunday.at('14:50').do(generate_report)

def calculate_time_until_next_event():
    now_utc = datetime.now(pytz.utc)
    next_event = min(schedule.get_jobs(), key=lambda x: x.next_run)
    next_event_utc = next_event.next_run.astimezone(pytz.utc)
    time_until_next_event = next_event_utc - now_utc
    return max(timedelta(seconds=0), time_until_next_event - timedelta(minutes=30))

while True:
    time_until_next_event = calculate_time_until_next_event().total_seconds()
    time.sleep(time_until_next_event)
    schedule.run_pending()
    
